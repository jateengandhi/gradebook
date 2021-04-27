# -*- coding: utf-8 -*-
"""
Created on Mon Sep 23 08:19:23 2019

@author: jatee
"""

import openpyxl
import datetime
import os.path
import json
from json import JSONEncoder # For serializing class


class GradeBook(object):
    '''
    This is a parent class, inherited by sub classes 'Excel' and 'Student'
    '''
    def __init__(self, course):

        # Course, semester, and year will be the folder names in the file path

        self.subject = 'subject'
        self.course = course  # 4 digit number, e.g. 1107
        self.semester = 'semester'  # See setSemester
        self.year = 2021

    def setSubject(self, subject):
        # Assumes subject is a string
        self.subject = subject

    def getSubject(self):
        return self.subject

    def setCourse(self, course):
        # Assumes course is an int, e.g. 2126        
        self.course = course 

    def getCourse(self):
        return self.course

    def setSemester(self, semester):

        # Assumes semester is an alphanumerical string:
        # F19 -> Fall 2021, S19 -> Spring 2021, Su19 -> Summer 2021
        self.semester = semester 


    def getSemester(self):
        return self.semester

    def setYear(self, year):
        self.year = year # e.g. 2021

    def getYear(self):
        return self.year

    def __str__(self):
        return self.subject + str(self.course) +', ' + self.semester

    def __repr__(self):
        return 'GradeBook('+str(self.course)+')'

class Excel(GradeBook):
    ''' 
    For reading and editing an excel file 
    '''

    def __init__(self, course, filename):

        # Parent class is GradeBook
        GradeBook.__init__(self, course)

        # filename.xlsx
        self.filename = filename

        # Borrows attributes from parent to set file path of .xlsx file
        self.path = 'pathname'

        # Open the .xlsx file for editing
        self.wb = openpyxl.load_workbook(self.path)
        self.ws = self.wb.active
        
        # Grab first sheet for editing. For more info see 'chooseSheet' below 
        self.chooseSheet()

    def chooseSheet(self, sheetInd = 0): 
        
        # This method has to be 'activated(!)' before working on the sheet
        self.Ind = sheetInd  # Default is 0
        self.wb.active = sheetInd  # Choose sheet from .xlsx file for writing.
        self.ws = self.wb.active  # Grab the active worksheet

    def __repr__(self):
        return 'Excel('+str(self.course)+')'  

    def setFilename(self, filename):
        # Change filename from original input
        self.filename = filename
        self.path = 'pathname'
        self.wb = openpyxl.load_workbook(self.path)
        self.chooseSheet()

    def getFilename(self):
        return self.filename

    def setPath(self, path):
        # Change default file path
        self.path = path
        self.wb = openpyxl.load_workbook(self.path)
        self.chooseSheet()

    def getPath(self):
        return self.path

    def getSheetNames(self):
        return self.wb.sheetnames

    def getSheetTitle(self): 
        # Title of the active sheet
        return self.ws.title

    def nRow(self):
        return self.ws.max_row

    def nCol(self):
        return self.ws.max_column

    def getHeaders(self):
        # Returns top cell from each column
        return list(col[0].value for col in self.ws.iter_cols(1,self.nCol(),1,1))

    def cellValue(self, row, column):
        # Returns data from selected cell
        # Note: Indexing begins at 1
        return self.ws.cell(row, column).value

    def chooseCell(self, row, column):
        # Select cell for editing
        # Note: Indexing begins at 1
        return self.ws.cell(row, column)

    def newSheet(self, title = 'newSheet'):
        # Copies the bio, which is the first 3 columns 
        # from the active sheet and pastes to a new sheet
        bio = {}
        n = 1
        upperLimit = 'C' + str(self.nRow())  # last row of 3rd column
        for row in self.ws['A1' : upperLimit]:  # Select range
            bio[n] = []
            for cell in row:
                bio[n].append(cell.value)
            n += 1

        self.wb.create_sheet(title = title, index = self.Ind+1)
        self.chooseSheet(self.Ind+1)
        for key in bio:
            for n, val in zip(range(1,4), bio[key]):
                self.ws.cell(key, n).value = val

    def copyCells(self, row_min, row_max, col_min, col_max):
        # Copy cells row wise. Creates a dictionary of rows as key, 
        # and corresponding columns in each row as a list of values
        self.all_cells = {}
        n = row_min
        for row in self.ws.iter_rows(min_row = row_min, max_row = row_max,
                                    min_col = col_min, max_col = col_max):
            self.all_cells[n] = []
            for cell in row:
                self.all_cells[n].append(cell.value)
            n += 1

    def pasteCells(self, sheet, row_min, row_max, col_min, col_max):
        # Overwrites cells without a warning
        self.chooseSheet(sheetInd = sheet)
        for key in self.all_cells:
            for col, val in zip(range(col_min,col_max+1), self.all_cells[key]):
                self.ws.cell(key, col).value = val

    def IDtoIndex(self, id_number):
        # Ientifies the row index for an ID number
        for row in range(1, self.nRow()+1):
            if self.cellValue(row, 1) == id_number:
                return row

    def TasktoIndex(self, task):
        # Identifies the index (row,column) for a given task
        for col in range(1,self.nCol()+1):
            if self.cellValue(1,col) == task:
                return col

    def writeCell(self, row, col, cel_val):
        # Writes data (score or Attendance) in a cell with 
        # the row number identified using IDtoIndex (see above)
        # and the column number identified using TasktoIndex (see above)
        self.chooseCell(row, col).value = cel_val


class Student(GradeBook):
    '''Each student's data is stored as a dict'''
    def __init__(self, course, Last, First):
        GradeBook.__init__(self, course)
        self.bio = {'Last': Last, 'First': First}
        self.Attendance = {}
        self.Quiz = {}
        self.Exam = {}
        self.Lab = {}
        self.aggregate = {}

    def setName(self, Last, First):
        self.bio['Last'] = Last
        self.bio['First'] = First

    def getLast(self):
        return self.bio['Last']
    
    def getFirst(self):
        return self.bio['First']
    
    def getFullName(self):
        return self.bio['Last'] + ', '+ self.bio['First']

    def __str__(self):
        return self.bio['Last'] + ', '+ self.bio['First']

    def __repr__(self):
        return 'Student(' + self.bio['Last'] + ', '+ self.bio['First'] + ')'

    def markAttend(self, month, day, year = 2021):
        date = datetime.date(year, month, day)
        mark = str(input('If '+'\033[1m' + self.bio['Last'] + ', '+ self.bio['First']
                        + '\033[0m' + ' is present enter P, else A: '))
        self.Attendance[date] = mark
        
    def getAttend(self):
        return self.Attendance
    
    def task(self, tsk):
        tsk_dict = {'Exam': self.Exam, 'Quiz': self.Quiz, 'Lab': self.Lab}
        return tsk_dict[tsk]
    
class GBEncoder(JSONEncoder):
    '''Serialize dictionary'''
    def default(self, o):
        return o.__dict__     
# scoring Excel and Student sub-classes

# Create an Excel instance for scores
# The first three columns of the excel file are Student ID, Last Name, First Name. 
# First row has those three headings, and subsequently task name and attendance will be added.
score = Excel('course number', 'file.xlsx')

# Create Gradebook.json if it does not exist
if os.path.isfile('Gradebook.json') == False:
    # Dictionary to combine both Excel and Student classes
    GB = {}

    # Keys in the dictionary will be the ID provided by the institute, 
    # and the value will be instance of class Student
    for row in range(2, score.nRow()+1):
        GB[score.cellValue(row,1)] = Student(score.getCourse(), 
                                            score.cellValue(row, 2), 
                                            score.cellValue(row, 3))

    # Serialization: Write the GB dict to a json file and save
    with open('Gradebook.json', 'w') as write_file:
        json.dump(GB, write_file, sort_keys=True, indent=4, cls=GBEncoder)


# Dictionary to combine both Excel and Student classes
GB = {}

 # Keys in the dictionary will be the ID provided by the institute, 
 # and the value will be instance of class Student
for row in range(2, score.nRow()+1):
     GB[score.cellValue(row,1)] = Student(score.getCourse(), 
                                           score.cellValue(row, 2), 
                                           score.cellValue(row, 3))


# Create a dictionary that lists students' last names and their ID. 'A' -> 'Anderson'
# Useful for seraching by last name.
def getNames():
    names = {}    
    for key in GB.keys():
        if GB[key].getLast()[0] not in names.keys():  # Search for the first letter of the last name as a key
            names[GB[key].getLast()[0]] = [(GB[key].getLast(), GB[key].getFirst(), key)]
        else:
            names[GB[key].getLast()[0]].append((GB[key].getLast(), GB[key].getFirst(), key))
            
    return names


def getIDs():
    # Returns a list of IDs
    return list(item for item in GB.keys())
     
        
def LastToID(names):
    # Using the input for the first letter of the last name, 
    # matching last names are listed, then selected last name's ID is returned
    letter = str(input('Enter the first letter of the last name: ')).upper()
    if letter not in names.keys():
        redo = str(input('No name found. Try another letter? (y/n): '))
        if redo == 'y':
            return LastToID(names)
        else:
            return None
    else:
        for i in range(len(names[letter])):
            response = str(input('Do you mean '+ str(names[letter][i][:-1]) + ' ? (y/n): '))
            if response == 'y':
                print('ID: '+str(names[letter][i][-1]))
                return names[letter][i][-1]
        redo = str(input('No name found. Try another letter? (y/n): '))
        if redo == 'y':
            return LastToID(names)
        else:
            return None

def selectTask():
    # To choose one of the tasks from Quiz, Lab, or Exam
    tsk = str(input('Enter task as either Quiz, Lab, or Exam:'))
    if tsk not in ('Exam', 'Quiz', 'Lab'):
        print('Invalid entry')
        return selectTask()
    return tsk

def enterGrade(month, day, year = 2021):
    # Gather info about task and create lists and dictionaries needed for grading

    # Create dictionary such as 'A' -> 'Anderson'
    names = getNames()

    # List of IDs  
    IDs = getIDs()

    # Day of the task
    date = datetime.date(year, month, day)
    date_str = date.__str__() 
    
    # Exam, score, assignment, or HW
    tsk = selectTask()

    # e.g. 1st or 2nd Exam
    task_number = int(input('Enter ' + tsk + ' number: '))

    # Enter maximum score for the task
    try:
        mx = float(input('Enter max grade possible: '))
    except ValueError:
        mx = float(input('Invalid entry. Please re-enter max grade possible: '))

    # Select the first empty column for writing grades in the score.xlsx
    grade_idx = score.nCol() + 1

    # Select the second empty column for entering Attendance
    attend_idx = score.nCol() + 2

    # Heading for the grade column
    score.writeCell(1, grade_idx, str(tsk) + '-' + str(task_number) + ' (' + str(mx) + ')')

    # Heading for the Attendance column
    score.writeCell(1, attend_idx, str(date))

    def Grading(IDs, tsk, task_number):
        # Enter grade and attendance for a student

        # Search ID from studet's last name
        ID = LastToID(names)

        if ID in IDs:

            # Add score for the selected task
            try:
                pts = float(input('Enter score for ' + str(GB[ID]) +': '))
            except ValueError:
                pts = float(input('Score needs to be input as a number. Please re-enter score for ' + str(GB[ID]) +': ' ))
                    
            # Record score in the task dictionary
            GB[ID].task(tsk)[task_number] = {'date' : date_str, 'max': mx, 'score': pts}

            # Record Attendance in the Attendance dictionary
            GB[ID].Attendance[date_str] = {str(tsk) + '-' + str(task_number): 'p'}

            row_idx = score.IDtoIndex(ID)
            score.writeCell(row_idx, grade_idx, pts)
            score.writeCell(row_idx, attend_idx, 'P')
      
            # Remove student's name from the dictionary 'names' 
            for item in names[str(GB[ID])[0]]:
                if item[-1] == ID:
                    names[str(GB[ID])[0]].remove(item)

            # Remove student's ID from the list 'IDs'        
            IDs.remove(ID)
            if len(IDs) == 0:
                print('')
                print('Grades have been entered for all students')
                print('')
                print('Make sure to re-run the score = Excel()')
                score.wb.save(score.getPath())
                with open('Gradebook.json', 'a') as append_file:
                    json.dump(GB, append_file, sort_keys=True, indent=4, cls=GBEncoder)
                return None

            return Grading(IDs, tsk, task_number)
        
        elif ID == None:
            print('ID not on the current roster')
            cont = str(input('Say y to continue grading or q to end grading: ' ))
            if cont == 'y':
                return Grading(IDs, tsk, task_number)
            elif cont == 'q':
                for ID in IDs:
                    GB[ID].task(tsk)[task_number] = {'date' : date_str, 'max': mx, 'score': 0}

                    # Record Attendance in the Attendance dictionary
                    GB[ID].Attendance[date_str] = {str(tsk) + '-' + str(task_number): 'A'}

                    row_idx = score.IDtoIndex(ID)
                    score.writeCell(row_idx, grade_idx, 0)
                    score.writeCell(row_idx, attend_idx, 'A')

                print('')
                print('Finished grading')
                print('')
                print('Make sure to re-run the score = Excel() if entering grades for additional tasks')
                score.wb.save(score.getPath())
                with open('Gradebook.json', 'a') as append_file:
                    json.dump(GB, append_file, sort_keys=True, indent=4, cls=GBEncoder)
                return None

    Grading(IDs, tsk, task_number)
